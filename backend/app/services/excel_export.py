import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from app.models import Product, Invoice, Transaction
from app.core.config import settings

class ExcelExportService:
    def __init__(self):
        self.export_path = settings.EXCEL_EXPORT_PATH
        os.makedirs(self.export_path, exist_ok=True)
    
    def export_products(self, db: Session) -> str:
        """Export all products to Excel"""
        products = db.query(Product).all()
        
        data = []
        for product in products:
            data.append({
                "SKU": product.sku,
                "Name": product.name,
                "Category": product.category,
                "Current Stock": product.current_stock,
                "Reorder Point": product.reorder_point,
                "Cost Price": product.cost_price,
                "Selling Price": product.selling_price,
                "Supplier": product.supplier,
                "Status": "Active" if product.is_active else "Inactive"
            })
        
        df = pd.DataFrame(data)
        filename = f"products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_path, filename)
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Products', index=False)
            
            # Format the Excel file
            workbook = writer.book
            worksheet = writer.sheets['Products']
            
            # Style the header
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Adjust column widths
            for idx, column in enumerate(df.columns, 1):
                column_letter = get_column_letter(idx)
                worksheet.column_dimensions[column_letter].width = max(12, len(column) + 2)
        
        return filepath
    
    def export_invoices(self, db: Session, status: str = None) -> str:
        """Export invoices to Excel"""
        query = db.query(Invoice)
        if status:
            query = query.filter(Invoice.status == status)
        
        invoices = query.all()
        
        data = []
        for invoice in invoices:
            data.append({
                "Invoice Number": invoice.invoice_number,
                "Customer": invoice.customer_name,
                "Issue Date": invoice.issue_date,
                "Due Date": invoice.due_date,
                "Status": invoice.status.value if invoice.status else "",
                "Subtotal": invoice.subtotal,
                "Tax": invoice.tax_amount,
                "Discount": invoice.discount_amount,
                "Total": invoice.total_amount,
                "Paid": invoice.paid_amount,
                "Balance": invoice.total_amount - invoice.paid_amount
            })
        
        df = pd.DataFrame(data)
        filename = f"invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_path, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Invoices', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Invoices']
            
            # Style the header
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Format currency columns
            currency_columns = ['Subtotal', 'Tax', 'Discount', 'Total', 'Paid', 'Balance']
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for idx, cell in enumerate(row):
                    if df.columns[idx] in currency_columns:
                        cell.number_format = '$#,##0.00'
            
            # Adjust column widths
            for idx, column in enumerate(df.columns, 1):
                column_letter = get_column_letter(idx)
                worksheet.column_dimensions[column_letter].width = max(12, len(column) + 2)
        
        return filepath
    
    def export_inventory_report(self, db: Session) -> str:
        """Export comprehensive inventory report"""
        products = db.query(Product).filter(Product.is_active == True).all()
        
        report_data = []
        total_value = 0
        low_stock_items = 0
        
        for product in products:
            stock_value = product.current_stock * product.cost_price
            total_value += stock_value
            
            if product.current_stock <= product.reorder_point:
                low_stock_items += 1
                stock_status = "Low Stock"
            elif product.current_stock == 0:
                stock_status = "Out of Stock"
            else:
                stock_status = "In Stock"
            
            report_data.append({
                "SKU": product.sku,
                "Product Name": product.name,
                "Category": product.category,
                "Current Stock": product.current_stock,
                "Min Stock": product.min_stock_level,
                "Max Stock": product.max_stock_level,
                "Reorder Point": product.reorder_point,
                "Stock Status": stock_status,
                "Unit Cost": product.cost_price,
                "Stock Value": stock_value,
                "Supplier": product.supplier
            })
        
        df = pd.DataFrame(report_data)
        
        # Create summary sheet
        summary_data = {
            "Metric": ["Total Products", "Total Stock Value", "Low Stock Items", "Out of Stock Items"],
            "Value": [
                len(products),
                f"${total_value:,.2f}",
                low_stock_items,
                len([p for p in products if p.current_stock == 0])
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_path, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write summary sheet
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Write detailed report
            df.to_excel(writer, sheet_name='Inventory Details', index=False)
            
            # Format both sheets
            for sheet_name in ['Summary', 'Inventory Details']:
                worksheet = writer.sheets[sheet_name]
                
                # Style headers
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Highlight low stock items in the details sheet
            details_sheet = writer.sheets['Inventory Details']
            for row in details_sheet.iter_rows(min_row=2, max_row=details_sheet.max_row):
                if row[7].value == "Low Stock":  # Stock Status column
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                elif row[7].value == "Out of Stock":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF")
        
        return filepath

excel_service = ExcelExportService()
